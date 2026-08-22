# Sinh file Excel mẫu để đội vận hành nhập sản phẩm đúng chuẩn.
# Chạy: python scripts/make-product-template.py
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUT = 'Hop-thu-den/MAU-NHAP-SAN-PHAM-KTD.xlsx'

# Lấy thẳng danh sách hãng và danh mục từ dữ liệu website để file mẫu không
# bao giờ lệch với thứ đang chạy thật.
src = io.open('lib/ktd-data.ts', encoding='utf-8').read()
brands = re.findall(r"\{ slug: '[a-z0-9-]+', name: '([^']+)'", src.split('export const CATEGORIES')[0])
cats = re.findall(r"\{ slug: '[a-z0-9-]+', name: '([^']+)', sub:", src)

BLUE = '006BB2'
LIGHT = 'EEF6FC'
GREY = 'F3F5F7'
RED = 'E30613'

thin = Side(style='thin', color='D4D9DE')
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ----------------------------------------------------------------- Hướng dẫn
ws = wb.active
ws.title = 'HƯỚNG DẪN'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 95

def head(row, text):
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(bold=True, size=13, color=BLUE)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

def line(row, label, text, warn=False):
    a = ws.cell(row=row, column=2, value=label)
    a.font = Font(bold=True, size=10, color=RED if warn else '111418')
    a.alignment = Alignment(vertical='top', wrap_text=True)
    b = ws.cell(row=row, column=3, value=text)
    b.font = Font(size=10, color=RED if warn else '3D444D')
    b.alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = 30

t = ws.cell(row=1, column=2, value='MẪU NHẬP SẢN PHẨM — CÔNG TY TNHH KIM THÀNH ĐÔNG')
t.font = Font(bold=True, size=16, color=BLUE)
ws.merge_cells('B1:C1')
ws.cell(row=2, column=2, value='Điền vào sheet "SẢN PHẨM". Sheet "DANH SÁCH CHỌN" là các giá trị hợp lệ.').font = Font(size=10, italic=True, color='6B747E')
ws.merge_cells('B2:C2')

r = 4
head(r, '⚠ QUAN TRỌNG NHẤT — ĐỌC TRƯỚC KHI NHẬP'); r += 1
line(r, 'Cột MÃ HÀNG phải để định dạng Text',
     'Excel tự đổi mã hàng thành số và LÀM MẤT SỐ 0 Ở CUỐI: 1031.50 thành 1031.5, '
     '10130610.02 thành 1.013061002E7. Cột này trong file mẫu đã khoá sẵn định dạng Text — '
     'đừng đổi lại. Nếu copy từ file khác, dùng Paste Special > Values.', warn=True); r += 1
line(r, 'Mỗi dòng là một mã hàng',
     'Một sản phẩm có nhiều mã (ví dụ SECUNORM MIZAR có 3 mã) thì tách thành nhiều dòng, '
     'mỗi dòng một mã riêng.'); r += 1
line(r, 'Không bỏ trống cột bắt buộc',
     'Các cột có dấu * là bắt buộc. Thiếu thì sản phẩm không lên web được.'); r += 2

head(r, '📝 VIẾT MÔ TẢ'); r += 1
line(r, 'MÔ TẢ NGẮN — 80 đến 160 ký tự',
     'Đây là câu khách đọc trên thẻ sản phẩm và trên kết quả Google. Phải là MỘT CÂU HOÀN CHỈNH, '
     'đủ nghĩa khi đứng riêng. Dưới 80 ký tự: hiện trọn trên thẻ. Trên 160: Google cắt bớt.'); r += 1
line(r, 'MÔ TẢ ĐẦY ĐỦ — không giới hạn',
     'Mỗi đoạn văn xuống một dòng (nhấn Alt+Enter trong ô). Hiện đầy đủ ở tab "Mô tả", '
     'không bị cắt.'); r += 1
line(r, 'Đừng lặp lại phần Ứng dụng',
     'Ứng dụng đã có cột riêng. Viết lại trong cột mô tả sẽ làm trang bị trùng nội dung.'); r += 1
line(r, 'ỨNG DỤNG — mỗi dòng một ý',
     'Nhấn Alt+Enter để xuống dòng trong ô. Không cần gõ dấu chấm đầu dòng.'); r += 2

head(r, '🔧 THÔNG SỐ KỸ THUẬT'); r += 1
line(r, 'Định dạng: Tên thông số: Giá trị',
     'Mỗi dòng một thông số, dùng dấu hai chấm. Ví dụ:\n'
     'Cơ chế an toàn: Lưỡi tự động thu\nĐộ sâu cắt: 18 mm\nChứng nhận: GS Certified'); ws.row_dimensions[r].height = 58; r += 1
line(r, 'Rất nên có',
     'Khách công nghiệp dựa vào bảng thông số để quyết định mua. Sản phẩm không có thông số '
     'thì tab này bị ẩn, trang mất một nửa giá trị.'); r += 2

head(r, '🖼 ẢNH SẢN PHẨM'); r += 1
line(r, 'Tối đa 5 ảnh mỗi mã hàng',
     'Ghi tên file, cách nhau bằng dấu phẩy. Ảnh đầu tiên là ảnh đại diện hiện trên thẻ sản phẩm.'); r += 1
line(r, 'Kích thước và nền',
     'Tối thiểu 1000 x 750 (tỉ lệ 4:3), nền trắng, sản phẩm cắt nét. Gửi ảnh gốc càng to càng tốt, '
     'bên kỹ thuật sẽ nén lại. Không cần tự nén.'); r += 1
line(r, 'Đặt tên file nên có mã hàng',
     'Ví dụ: secupro-martego-122001-1.jpg. Giúp ghép ảnh đúng sản phẩm, tránh nhầm lẫn.'); r += 1
line(r, 'Để ảnh trong thư mục riêng',
     'Cách nhanh hơn liệt kê từng tên file: mỗi sản phẩm một thư mục, đặt tên thư mục theo tên '
     'sản phẩm, rồi ghi tên thư mục vào cột ảnh. Bên kỹ thuật tự lấy hết ảnh trong đó.'); r += 2

head(r, '🔢 THỨ TỰ HIỂN THỊ'); r += 1
line(r, 'Cột NỔI BẬT',
     'Đánh dấu x thì sản phẩm hiện ở khối "Sản phẩm được quan tâm" trên trang chủ. '
     'Nên chọn khoảng 8 mã tiêu biểu.'); r += 1
line(r, 'Cột ƯU TIÊN HIỆN MẶC ĐỊNH',
     'Quyết định thứ tự sản phẩm khi khách vào trang "Sản phẩm" mà chưa lọc gì. '
     'Điền số: 1 hiện trước, 2 hiện sau... Để trống thì xếp xuống cuối. '
     'Dùng để đẩy hàng chủ lực lên đầu.'); r += 2

head(r, '📄 TÀI LIỆU'); r += 1
line(r, 'FILE PDF',
     'Datasheet hoặc catalog của mã hàng. Ghi tên file, để cùng thư mục ảnh.'); r += 1
line(r, 'LINK TRANG HÃNG',
     'Đường dẫn tới trang sản phẩm trên website của hãng, nếu có.'); r += 2

head(r, '✅ TỪ KHÓA TÌM KIẾM'); r += 1
line(r, 'Gồm cả tên gọi dân dã ngoài xưởng',
     'Đây là điểm mạnh của website: khách gõ "pa lang", "mui mai ca rem", "sung mai hoi" vẫn ra '
     'đúng sản phẩm. Hãy ghi cả những tên mà thợ máy thật sự dùng, cách nhau bằng dấu phẩy. '
     'Không cần gõ dấu tiếng Việt.'); ws.row_dimensions[r].height = 45

# ----------------------------------------------------------------- Sản phẩm
COLS = [
    ('STT', 6, '', ''),
    ('THƯƠNG HIỆU *', 20, 'Chọn từ danh sách', 'Martor'),
    ('DANH MỤC *', 30, 'Chọn từ danh sách', 'Dụng cụ an toàn'),
    ('TÊN SẢN PHẨM *', 34, 'Tối đa 60 ký tự', 'SECUNORM SMARTCUT MDP'),
    ('MÃ HÀNG *', 18, 'ĐỊNH DẠNG TEXT — giữ nguyên số 0 cuối', '110700.02'),
    ('DÒNG SẢN PHẨM', 18, 'Series', 'SECUNORM'),
    ('XUẤT XỨ', 14, '', 'Đức'),
    ('MÔ TẢ NGẮN *', 52, '80–160 ký tự, một câu hoàn chỉnh',
     'Dao an toàn MARTOR SECUNORM SMARTCUT MDP là dao cắt an toàn dùng một lần cho môi trường '
     'kiểm soát dị vật.'),
    ('MÔ TẢ ĐẦY ĐỦ *', 60, 'Mỗi đoạn một dòng (Alt+Enter)',
     'Thân dao bằng nhựa phát hiện kim loại (MDP).\nLưỡi tự động thu khi rời khỏi vật liệu cắt.'),
    ('ỨNG DỤNG', 44, 'Mỗi dòng một ý (Alt+Enter)',
     'Cắt bao bì đóng gói\nCắt màng nhựa và màng film\nNgành thực phẩm, dược phẩm'),
    ('THÔNG SỐ KỸ THUẬT', 44, 'Tên: Giá trị — mỗi dòng một thông số',
     'Cơ chế an toàn: Lưỡi tự động thu\nVật liệu: Nhựa MDP\nChứng nhận: GS Certified'),
    ('TỪ KHÓA TÌM KIẾM', 34, 'Cách nhau dấu phẩy, gồm tên dân dã',
     'dao an toan, dao roc giay, smartcut'),
    ('TÊN FILE ẢNH *', 40,
     CellRichText(
         TextBlock(InlineFont(sz=9, i=True, color='6B747E'), 'Tối đa 5 file, cách nhau dấu phẩy / '),
         TextBlock(InlineFont(sz=9, i=True, b=True, color=RED), 'hoặc bỏ vào 1 thư mục'),
     ),
     'smartcut-mdp-110700-1.jpg, smartcut-mdp-110700-2.jpg'),
    ('TÊN FILE PDF', 26, 'Datasheet nếu có', 'smartcut-mdp-110700.pdf'),
    ('LINK TRANG HÃNG', 34, 'Nếu có', 'https://martor.com/...'),
    ('NỔI BẬT', 12, 'x = hiện ở trang chủ', ''),
    ('ƯU TIÊN HIỆN MẶC ĐỊNH TRANG "SẢN PHẨM"', 22, 'Số nhỏ hiện trước. Trống = xếp cuối', '1'),
]

ws2 = wb.create_sheet('SẢN PHẨM')
ws2.freeze_panes = 'A3'

for i, (name, width, hint, example) in enumerate(COLS, start=1):
    L = get_column_letter(i)
    ws2.column_dimensions[L].width = width

    h = ws2.cell(row=1, column=i, value=name)
    h.fill = PatternFill('solid', fgColor=BLUE)
    h.font = Font(bold=True, color='FFFFFF', size=10)
    h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    h.border = box

    g = ws2.cell(row=2, column=i, value=hint)
    g.fill = PatternFill('solid', fgColor=LIGHT)
    if isinstance(hint, str):
        g.font = Font(size=9, italic=True, color=RED if 'TEXT' in hint else '6B747E')
    g.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    g.border = box

    e = ws2.cell(row=3, column=i, value=example)
    e.fill = PatternFill('solid', fgColor=GREY)
    e.font = Font(size=9, color='6B747E', italic=True)
    e.alignment = Alignment(vertical='top', wrap_text=True)
    e.border = box

ws2.cell(row=3, column=1, value='(ví dụ)')
ws2.row_dimensions[1].height = 34
ws2.row_dimensions[2].height = 28
ws2.row_dimensions[3].height = 74

# Cột mã hàng khoá định dạng Text cho 500 dòng — đây là lỗi hay gặp nhất
for row in range(3, 504):
    ws2.cell(row=row, column=5).number_format = '@'

# Danh sách chọn cho Thương hiệu và Danh mục
dv_b = DataValidation(type='list', formula1=f"'DANH SÁCH CHỌN'!$A$2:$A${len(brands)+1}", allow_blank=True)
dv_c = DataValidation(type='list', formula1=f"'DANH SÁCH CHỌN'!$C$2:$C${len(cats)+1}", allow_blank=True)
ws2.add_data_validation(dv_b)
ws2.add_data_validation(dv_c)
dv_b.add(f'B4:B503')
dv_c.add(f'C4:C503')

# ----------------------------------------------------------------- Danh sách chọn
ws3 = wb.create_sheet('DANH SÁCH CHỌN')
ws3.column_dimensions['A'].width = 26
ws3.column_dimensions['C'].width = 40
for col, title, values in ((1, f'THƯƠNG HIỆU ({len(brands)})', brands),
                           (3, f'DANH MỤC ({len(cats)})', cats)):
    h = ws3.cell(row=1, column=col, value=title)
    h.fill = PatternFill('solid', fgColor=BLUE)
    h.font = Font(bold=True, color='FFFFFF', size=10)
    for i, v in enumerate(values, start=2):
        c = ws3.cell(row=i, column=col, value=v)
        c.font = Font(size=10)
        c.border = box

wb.save(OUT)
print(f'đã tạo {OUT}')
print(f'  {len(COLS)} cột · {len(brands)} thương hiệu · {len(cats)} danh mục')
